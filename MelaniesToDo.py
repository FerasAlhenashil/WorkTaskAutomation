import os
from datetime import datetime, timedelta
import re
import warnings

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side


class Melaniestodo_today(object):
    project_indices = []
    ws_rows = []
    jobs = []
    matches = []
    todo_today = []
    todo_tomorrow = []
    todo_2days = []
    output_today = []
    output_tomorrow = []
    output_2days = []

    def __init__(self):
        pass

    def extract(self, ws):
        fini = self.get_fini(ws)
        for row in ws:
            for cell in row:
                if cell.value == 'Status':
                    self.project_indices.append(cell.row - 1)
        self.project_indices.append(fini)

        for row in ws.values:
            self.ws_rows.append(list(row))

        for i in range(len(self.project_indices) - 1):
            temp = []
            start = self.project_indices[i]
            end = self.project_indices[i + 1]
            for j in range(end - start):
                temp.append(self.ws_rows[start + j])
            self.jobs.append(temp)

    def match(self):
        for i in range(len(self.jobs)):
            for j in range(len(self.jobs[i])):
                try:
                    temp = re.findall("((?:[1-9]|1[0-2])/(?:\\d)*)", self.jobs[i][j][4])
                except TypeError:
                    continue
                if temp:
                    #record loops position to be the key for finding a task
                    temp.append([i, j])
                    self.matches.append(temp)

    def format_date(self):
        for i in range(len(self.matches)):
            for j in range(len(self.matches[i]) - 1):
                self.matches[i][j] += '/19'
                self.matches[i][j] = datetime.strptime(self.matches[i][j], "%m/%d/%y")
        for i in range(len(self.jobs)):
            for j in range(len(self.jobs[i])):
                try:
                    self.jobs[i][j][4] = self.jobs[i][j][4].split(",")
                except AttributeError:
                    continue

    def get_fini(self, ws):
        fini = ws.max_row
        if fini > 1:
            while ws.cell(row=fini, column=1).value is None:
                fini -= 1
        return fini

    def check(self):
        today = datetime.today() + relativedelta(days=0)
        for i in range(len(self.matches)):
            for j in range(len(self.matches[i])-1):
                if timedelta(hours=-24) < (self.matches[i][j] - today) < timedelta(hours=0):
                    temp = [self.matches[i][-1][0], self.matches[i][-1][1], j]
                    self.todo_today.append(temp)

        for i in range(len(self.matches)):
            for j in range(len(self.matches[i])-1):
                if today.weekday() < 3:
                    if timedelta(hours=-6) < (self.matches[i][j] - today) < timedelta(days=1):
                        temp = [self.matches[i][-1][0], self.matches[i][-1][1], j]
                        self.todo_tomorrow.append(temp)
                else:
                    if timedelta(hours=1) < (self.matches[i][j] - today) < timedelta(days=3):
                        temp = [self.matches[i][-1][0], self.matches[i][-1][1], j]
                        self.todo_tomorrow.append(temp)

        for i in range(len(self.matches)):
            for j in range(len(self.matches[i])-1):
                if today.weekday() < 3:
                    if timedelta(days=1) < (self.matches[i][j] - today) < timedelta(days=2):
                        temp = [self.matches[i][-1][0], self.matches[i][-1][1], j]
                        self.todo_2days.append(temp)
                elif today.weekday() == 3:
                    if timedelta(days=2) < (self.matches[i][j] - today) < timedelta(days=4):
                        temp = [self.matches[i][-1][0], self.matches[i][-1][1], j]
                        self.todo_2days.append(temp)
                else:
                    if timedelta(days=3) < (self.matches[i][j] - today) < timedelta(days=4):
                        temp = [self.matches[i][-1][0], self.matches[i][-1][1], j]
                        self.todo_2days.append(temp)

    def to_do(self):
        print(" ")
        for i in range(len(self.todo_today)):
            temp = [self.jobs[self.todo_today[i][0]][0][0], self.jobs[self.todo_today[i][0]][self.todo_today[i][1]][0],
                    self.jobs[self.todo_today[i][0]][self.todo_today[i][1]][4][self.todo_today[i][2]]]
            self.output_today.append(temp)

        for i in range(len(self.todo_tomorrow)):
            temp = [self.jobs[self.todo_tomorrow[i][0]][0][0],
                    self.jobs[self.todo_tomorrow[i][0]][self.todo_tomorrow[i][1]][0],
                    self.jobs[self.todo_tomorrow[i][0]][self.todo_tomorrow[i][1]][4][self.todo_tomorrow[i][2]]]
            self.output_tomorrow.append(temp)

        for i in range(len(self.todo_2days)):
            temp = [self.jobs[self.todo_2days[i][0]][0][0], self.jobs[self.todo_2days[i][0]][self.todo_2days[i][1]][0],
                    self.jobs[self.todo_2days[i][0]][self.todo_2days[i][1]][4][self.todo_2days[i][2]]]
            self.output_2days.append(temp)

        for i in range(len(self.output_today)):
            task = self.output_today[i][2]
            if task.endswith("PM Prep"):
                task += " \"Inputs due from PM\""
            elif task.endswith("Prep2"):
                task += " \"Due for Q&A\""
            elif task.endswith("Q&A"):
                task += " \"Feedback due from Cambridge - send to prep 3\""
            elif task.endswith("Prep3"):
                task += " \"Inputs for Format due from GK\""
            elif task.endswith("FQA"):
                task += " \"Due from FQA\""
            elif task.endswith("FQA/CEPR/PMTSGK") or task.endswith("CEPR/PMTSGK") or task.endswith("FQA/PMTSGK"):
                task += " \"Due from entry PMTSGK\""
            elif task.endswith("CE"):
                task += " \"Due from CE - send to entry or ICR\""
            elif task.endswith("CE/Upload for ICR") or task.endswith("CE/Upload for ICA"):
                task += " \"CE - send to ICR\""
            elif task.endswith("Finalize") or task.endswith("Delivery Prep"):
                task += " \"Due from delivery prep\""
            elif task.endswith("PDQA ") or task.endswith("PDQA"):
                task += " \"Due from PDQA\""
            elif task.endswith("Delivery") or task.endswith("FFCH"):
                task += " \"Due from FFCH\""
            elif task.endswith("GK Format"):
                task += " \"Formatted files due. Send to proofreading/Check Entry\""
            elif task.endswith("Format"):
                task += " \"Formatted files due. Send to proofreading/Check Entry\""
            elif task.endswith("PR/CE") or task.endswith("Proof/CE") or task.endswith("CE/PR") or \
                    task.endswith("CE/Proof") or task.endswith("CE (LING)") or task.endswith("CE + PR"):
                task += " \"Complete CE and verify linguistic feedbackhas been incorporated." \
                        " Send to entry or upload for ICR/ICA\""
            elif task.endswith("PR/CE/Upload for ICA") or task.endswith("PR/CE/Upload for ICR"):
                task += " \"Complete CE and possible LSO send to ICR/ICA upload\""
            elif task.endswith(" PMTSGK") or task.endswith("PM/TS CH") or task.endswith("TS + GK"):
                task += " \"TS Check and send to CEPR or Entry\""
            elif task.endswith("Entry ICA") or task.endswith("Entry ICR"):
                task += " \"Due from entry of ICR/ICA. Coordinate with PM for CE assignment.\""
            self.output_today[i][2] = task

        for i in range(len(self.output_tomorrow)):
            task = self.output_tomorrow[i][2]
            if task.endswith("Prep1"):
                task += " \"Send inputs to prep\""
            elif task.endswith("Xlation") or task.endswith("Xlation "):
                task += " \"Send to format\""
            elif task.endswith("Proof"):
                task += " \"Send formatted file to PM for Proof\""
            elif task.endswith("Enter PR"):
                task += " \"Send to Entry of Proof or skip to FQA\""
            elif task.endswith("FQA"):
                task += " \"Send to FQA\""
            elif task.endswith("CEPR"):
                task += " \"Verify checks are complete and notify PM that CEPR is ready for launch.\""
            elif task.endswith("FQA/CEPR/PMTSGK") or task.endswith("CEPR/PMTSGK") or task.endswith("FQA/PMTSGK"):
                task += " \" Send to entry\""
            elif task.endswith("CE/Upload for ICR") or task.endswith("CE/Upload for ICA"):
                task += " \"Send to entry of CE\""
            elif task.endswith("GK Format") or task.endswith("GK Prep/FO"):
                task += " \"Send inputs to GK format\""
            elif task.endswith("Format"):
                task += " \"Send inputs to format\""
            elif task.endswith(" PMTSGK") or task.endswith("PM/TS CH") or task.endswith("TS + GK"):
                task += " \"Launch PMTSGK\""
            self.output_tomorrow[i][2] = task

        for i in range(len(self.output_2days)):
            task = self.output_2days[i][2]
            if task.endswith("Xlation") or task.endswith("Xlation "):
                task += " \"translations due from PM\""
            self.output_2days[i][2] = task

    def output(self, todo_ws, wb):
        thin = Side(border_style="thin", color="000000")
        ft = Font(name='Calibri', size=16)
        top_style = NamedStyle(name="top_style", font=ft, alignment=Alignment(horizontal="center", vertical="center"),
                               border=Border(top=thin, left=thin, right=thin, bottom=thin))
        body1 = NamedStyle(name="body1", font=ft, border=Border(top=thin, left=thin, right=
            thin, bottom=thin), alignment=Alignment(horizontal="left", vertical="top"))
        try:
            wb.add_named_style(top_style)
            wb.add_named_style(body1)
        except ValueError:
            pass

        for i, j in enumerate(self.output_today[:-1]):
            fini = self.get_fini(todo_ws)
            todo_ws.row_dimensions[fini + 1].height = 70
            todo_ws.column_dimensions['A'].width = 70
            todo_ws.column_dimensions['B'].width = 100
            todo_ws.row_dimensions[fini + 2].height = 50
            cell1 = todo_ws.cell(row=fini + 1, column=1)
            cell2 = todo_ws.cell(row=fini + 2, column=1)
            cell3 = todo_ws.cell(row=fini + 2, column=2)
            cell4 = todo_ws.cell(row=fini + 1, column=1)
            cell5 = todo_ws.cell(row=fini + 1, column=2)
            cell1.style, cell2.style, cell3.style, cell4.style, cell5.style = 'top_style', 'body1', 'body1', 'body1', \
                                                                              'body1'
            current = j[0]
            art_work = j[1]
            task = j[2]
            previous = self.output_today[i-1][0]
            #if the project name is the same print it at least once
            if i == 0 and current == previous:
                cell1.value = current
            #tasks longer than 20 contain messages
            if current != previous and len(task) > 20:
                cell1.value = current
                cell2.value = art_work
                cell3.value = task
            elif len(task) > 20:
                cell4.value = art_work
                cell5.value = task

    def api(self):
        os.chdir('C:\\Users\\Feras\\Documents\\Work_Task_Automaiton')
        tracker = 'Melanies_Project_Tracker_MW (2) - Copy'
        tracker_file = tracker + '.xlsx'
        wb = load_workbook(tracker_file)
        ws = wb.active
        todo = 'Todo_lists'
        todo_file = todo + '.xlsx'
        todo_wb = load_workbook(todo_file)
        todo_ws = todo_wb.active
        self.extract(ws)
        self.match()
        self.format_date()
        self.check()
        self.to_do()
        self.output(todo_ws, todo_wb)
        try:
            todo_wb.save(todo_file)
        except PermissionError:
            input("Unable to save to file, make sure it's not open\n\nPress any key to exit")


def main():
    obj = Melaniestodo_today()
    obj.api()


if __name__ == '__main__':
    main()
