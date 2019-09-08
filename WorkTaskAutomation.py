import os
import warnings
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
import tkinter as tk
from tkinter import *
from tkinter import messagebox
import tkinter.font


class GUI(tk.Frame):
    def __init__(self, master):
        self.window = master
        tk.Frame.__init__(self, master)
        self.window.title("Tracker Generator")
        w = 354
        h = 150
        self.window.minsize(width=w, height=h)
        ws = self.window.winfo_screenwidth()
        hs = self.window.winfo_screenheight()
        x = (ws/2) - (w/2) - h
        y = (hs / 2) - (h / 2) - h
        self.window.geometry('%dx%d+%d+%d' % (w, h, x, y))
        tk_rgb = "#%02x%02x%02x" % (240, 212, 252)
        self.window.configure(bg=tk_rgb)
        childObj = WorkTaskAutomation(self)
        self.window.mainloop()


class WorkTaskAutomation(object):
    titles = []
    art_works = []
    schedule = []
    schedule_dates = []

    def __init__(self, gui):
        self.gui = gui
        helv36 = tkinter.font.Font(family="Comic Sans MS", size=11, weight="bold")
        tk_rgb = "#%02x%02x%02x" % (240, 212, 252)
        tkb_rgb = "#%02x%02x%02x" % (252, 169, 191)
        schedule_label = Label(gui.window, text="Schedule file:", font=helv36, bg=tk_rgb, pady=3)
        schedule_label.grid(row=0, column=0)
        schedule_file = StringVar()
        self.schedule_E = Entry(gui.window, textvariable=schedule_file, width=35)
        self.schedule_E.grid(row=0, column=1)
        tracker_label1 = Label(gui.window, text="Tracker to load:", font=helv36, bg=tk_rgb, pady=3)
        tracker_label1.grid(row=1, column=0)
        tracker_file1 = StringVar()
        self.tracker_E1 = Entry(gui.window, textvariable=tracker_file1, width=35)
        self.tracker_E1.grid(row=1, column=1)
        tracker_label2 = Label(gui.window, text="Tracker to save:", font=helv36, bg=tk_rgb, pady=3)
        tracker_label2.grid(row=2, column=0)
        tracker_file2 = StringVar()
        self.tracker_E2 = Entry(gui.window, textvariable=tracker_file2, width=35)
        self.tracker_E2.grid(row=2, column=1)
        self.button = tk.Button(gui.window, text="Generate  tracker",font=helv36, command=self.API,
                                relief=RIDGE, bg=tkb_rgb, padx=30).grid(row=3, column=1, pady=(10, 0))

    def ws_titles(self, ws):
        for row in ws:
            for cell in row:
                if cell.column is 7 and cell.value is None:
                    cell = cell.offset(0, 1)
                    if cell.value is not None:
                        self.titles.append(cell.value)

    def ws_art_works(self, ws):
        for row in ws.values:
            for value in row:
                if value == 'Melanie':
                    self.art_works.append(list(row))
        for i in range(len(self.art_works)):
            for j in range(len(self.art_works[i])):
                if type(self.art_works[i][j]) is not str:
                    self.art_works[i][j] = ''

    def ws_project_starting_indices(self):
        project_starting_indices = [0]
        j = 0
        for i in range(len(self.art_works)):
            if self.art_works[j][3] != self.art_works[i][3]:
                project_starting_indices.append(i)
                j = i
        return project_starting_indices

    def ws_schedule(self):
        for i in range(len(self.art_works)):
            try:
                start = [i for i, s in enumerate(self.art_works[i]) if 'DV' in s or 'PM ' in s]
                end = [i for i, s in enumerate(self.art_works[i]) if 'ICR' in s or 'ICA' in s]
                self.schedule.append(self.art_works[i][start[0]:end[0]+1])
            except ValueError:
                print('project might not contain DV, PM, ICR or ICA tasks')

    def ws_schedule_dates(self, ws):
        dates = []
        for row in ws.values:
            for value in row:
                dates.append(value)
        for i in range(len(self.art_works)):
            try:
                start = [i for i, s in enumerate(self.art_works[i]) if 'DV' in s or 'PM ' in s]
                end = [i for i, s in enumerate(self.art_works[i]) if 'ICR' in s or 'ICA' in s]
                self.schedule_dates.append(dates[start[0]:end[0]+1])
            except ValueError:
                pass

    def ws_tracker_schedule(self):
        tracker_schedule = []
        for i in range(len(self.schedule_dates)):
            cell_schedule = []
            for j in range(len(self.schedule_dates[i])):
                try:
                    schedule_str = " " + str(self.schedule_dates[i][j].month) + "/" \
                                   + str(self.schedule_dates[i][j].day)+ " " + str(self.schedule[i][j])
                    if (self.schedule[i][j]) != '':
                        cell_schedule.append(schedule_str)
                except AttributeError:
                    print('An error with schedule dates')
            tracker_schedule.append(cell_schedule)
        return tracker_schedule

    def get_fini(self, tracker_ws):
        fini = tracker_ws.max_row
        while tracker_ws.cell(row=fini, column=1).value is None:
            fini -= 1
        return fini

    def ws_append(self, tracker_ws, tracker_schedule, project_starting_indices, tracker_wb):
        fini = self.get_fini(tracker_ws)
        try:
            tracker_ws.unmerge_cells(start_row=fini, start_column=1, end_row=fini, end_column=6)
        except ValueError:
            print('cells are not merged')
        thin = Side(border_style="thin", color="000000")
        ft = Font(name='Calibri', size=16)
        top_style = NamedStyle(name="top_style", fill=PatternFill(patternType='solid', fill_type='solid', fgColor=
            "c5d9f1"), font=ft, alignment=Alignment(horizontal="center", vertical="center"),
                               border=Border(top=thin, left=thin, right=thin, bottom=thin))
        body1 = NamedStyle(name="body1", font=ft, border=Border(top=thin, left=thin, right=
            thin, bottom=thin), alignment=Alignment(horizontal="left", vertical="top"))
        try:
            tracker_wb.add_named_style(top_style)
            tracker_wb.add_named_style(body1)
        except ValueError:
            pass
        for i in range(len(self.titles)):
            fini = self.get_fini(tracker_ws)
            tracker_ws.insert_rows(fini)
            try:
                self.tracker_header(tracker_ws, i, project_starting_indices[i])
                self.tracker_body(tracker_ws, tracker_schedule, project_starting_indices[i])
            except IndexError:
                print('There was an issue with the arrays that caused an error')

    def tracker_header(self, tracker_ws, i, project_starting_indices):
        fini = self.get_fini(tracker_ws)
        tracker_ws.row_dimensions[fini-1].height = 100
        cell1 = tracker_ws.cell(row=fini-1, column=1)
        cell2 = tracker_ws.cell(row=fini-1, column=2)
        cell3 = tracker_ws.cell(row=fini-1, column=3)
        cell4 = tracker_ws.cell(row=fini-1, column=4)
        cell5 = tracker_ws.cell(row=fini-1, column=5)
        cell6 = tracker_ws.cell(row=fini-1, column=6)

        cell1.value = str(self.titles[i]) + '   ' + '(' + str(self.art_works[project_starting_indices][3]) + ')'
        cell2.value = 'GK'
        cell3.value = 'DTP'
        cell4.value = 'Status'
        cell5.value = 'Schedule'
        cell6.value = 'Notes'
        cell1.style, cell2.style, cell3.style, cell4.style, cell5.style, cell6.style = 'top_style', 'top_style',\
                                                                                       'top_style', 'top_style',\
                                                                                       'top_style', 'top_style'

    def tracker_body(self, tracker_ws, tracker_schedule, project_starting_indices):
        i = project_starting_indices
        fini = self.get_fini(tracker_ws)
        while self.art_works[project_starting_indices][3] == self.art_works[i][3]:
            tracker_ws.insert_rows(fini)
            fini = self.get_fini(tracker_ws)
            tracker_ws.row_dimensions[fini - 1].height = 70
            cell1 = tracker_ws.cell(row=fini - 1, column=1)
            cell2 = tracker_ws.cell(row=fini - 1, column=2)
            cell3 = tracker_ws.cell(row=fini - 1, column=3)
            cell5 = tracker_ws.cell(row=fini - 1, column=5)
            cell1.value = self.art_works[i][6] + '   ' + self.art_works[i][7]
            cell2.value = self.art_works[i][1]
            cell3.value = self.art_works[i][2]
            cell5.value = ','.join(tracker_schedule[i])
            cell1.style, cell2.style, cell3.style, cell5.style = 'body1', 'body1', 'body1', 'body1'
            i += 1
            if i > len(self.art_works)-1:
                break

    def API(self):
        #'C:\\Users\\Feras\\Documents\\Work_Task_Automaiton'
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            schedule_file = self.schedule_E.get() + '.xlsm'
            try:
                os.chdir('R:\\LifeScan\\_Lifescan_General\\20_Personal_Folders\\Melanie\\Schedules')
                wb = load_workbook(schedule_file)
            except FileNotFoundError:
                messagebox.showinfo("Error", "Unable to find the schedule file. Make sure it exists at the expected location")
                quit()
            ws = wb.active
            tracker_file = self.tracker_E1.get() + '.xlsx'
            try:
                tracker_wb = load_workbook(tracker_file)
            except FileNotFoundError:
                messagebox.showinfo("Error", "Unable to find the tracker file. Make sure it exists at the expected location")
                quit()
            tracker_ws = tracker_wb.active
            self.ws_art_works(ws)
            project_starting_indices = self.ws_project_starting_indices()
            self.ws_titles(ws)
            self.ws_schedule()
            self.ws_schedule_dates(ws)
            tracker_schedule = self.ws_tracker_schedule()
            self.ws_append(tracker_ws, tracker_schedule, project_starting_indices, tracker_wb)
            fini = self.get_fini(tracker_ws)
            tracker_ws.merge_cells(start_row=fini, start_column=1, end_row=fini, end_column=6)
            tracker_file2 = self.tracker_E2.get() + '.xlsx'
            try:
                tracker_wb.save(tracker_file2)
            except PermissionError:
                messagebox.showinfo("Error", "\nUnable to save to the tracker. Please make sure the file isn\'t open")
                quit()


def main():
    window = tk.Tk()
    obj = GUI(window)
    window.mainloop()


if __name__ == '__main__':
    main()
